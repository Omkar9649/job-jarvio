"""Generate Job Jarvios project plan Excel workbook."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT = "Job_Jarvios_Project_Plan.xlsx"

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header_row(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_rows(ws, start_row, end_row, col_count):
    for row in range(start_row, end_row + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def auto_width(ws, min_width=12, max_width=45):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[letter].width = min(max(length + 2, min_width), max_width)


def sheet_overview(wb):
    ws = wb.active
    ws.title = "Project Overview"

    rows = [
        ["Field", "Details"],
        ["Project Name", "Job Jarvios - Personal AI Career Assistant"],
        ["Owner", "Omkar (Personal Use)"],
        ["Goal", "Automatically find and rank relevant jobs from target companies based on resume/skills"],
        ["Start Date", "June 2026"],
        ["Target MVP", "4 Weeks"],
        ["Budget", "Free (self-hosted: Ollama, n8n, Qdrant, SQLite/Postgres)"],
        ["Primary Stack", "Node.js, Python, n8n, Ollama, Qdrant, Cheerio/BeautifulSoup"],
        ["Notification", "Email (Gmail) / Telegram (Phase 1)"],
        ["Job Sources (Phase 1)", "Company career pages from AmbitionBox company list"],
        ["Job Sources (Phase 2)", "Naukri/LinkedIn job alert emails via IMAP"],
        ["Resume Value", "Demonstrates scraping, automation, RAG, vector search, local LLM"],
        ["GitHub Repo", "job-jarvios"],
        ["Success Metrics", "Daily top 5-10 relevant jobs; reduced manual search time; interview callbacks"],
    ]

    for r, row in enumerate(rows, 1):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    style_header_row(ws, 1, 2)
    style_data_rows(ws, 2, len(rows), 2)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 70


def sheet_phases(wb):
    ws = wb.create_sheet("Phase Plan")
    headers = ["Phase", "Week", "Focus", "Deliverables", "Status"]
    data = [
        ["Phase 0 - Setup", "Week 0", "Environment & repo setup", "Docker (n8n, Ollama, Qdrant), GitHub repo, folder structure", "In Progress"],
        ["Phase 1 - Data Collection", "Week 1", "Company & job ingestion", "AmbitionBox company CSV, career URL finder, job scraper, SQLite DB", "Planned"],
        ["Phase 2 - AI Matching", "Week 2", "Resume + vector search", "Resume parser, Ollama embeddings, Qdrant index, similarity match", "Planned"],
        ["Phase 3 - RAG Ranking", "Week 3", "LLM reranking", "Ollama prompt for top job ranking with match reason & missing skills", "Planned"],
        ["Phase 4 - Automation", "Week 3-4", "n8n workflows", "Daily cron: scrape -> match -> notify pipeline", "Planned"],
        ["Phase 5 - Alerts", "Week 4", "Notifications", "Email/Telegram daily digest of top matched jobs", "Planned"],
        ["Phase 6 - Portfolio", "Week 4", "Documentation", "README, architecture diagram, demo video, resume bullets", "Planned"],
        ["Phase 7 - Enhance", "Later", "Email alert ingestion", "Parse Naukri/LinkedIn alert emails into same pipeline", "Backlog"],
        ["Phase 8 - UI", "Later", "Dashboard", "Simple web UI: matches, applied jobs, saved companies", "Backlog"],
    ]

    ws.append(headers)
    for row in data:
        ws.append(row)

    style_header_row(ws, 1, len(headers))
    style_data_rows(ws, 2, len(data) + 1, len(headers))
    auto_width(ws)


def sheet_tasks(wb):
    ws = wb.create_sheet("Task Breakdown")
    headers = [
        "ID", "Phase", "Task", "Description", "Tech/Tools",
        "Priority", "Est. Hours", "Week", "Status", "Dependencies",
    ]
    data = [
        ["T01", "Phase 0", "Create project folder structure", "scraper/, matcher/, n8n/, docs/, data/", "Git", "High", 2, 0, "In Progress", "-"],
        ["T02", "Phase 0", "Docker Compose setup", "n8n + Ollama + Qdrant containers", "Docker", "High", 4, 0, "Planned", "T01"],
        ["T03", "Phase 0", "Install Ollama models", "Pull llama3.2 and nomic-embed-text", "Ollama", "High", 1, 0, "Planned", "T02"],
        ["T04", "Phase 1", "Scrape AmbitionBox companies", "Filtered IT companies Mumbai/Pune - CSV output", "Node.js/Cheerio", "High", 4, 1, "Done", "-"],
        ["T05", "Phase 1", "Design database schema", "companies, jobs, resume_chunks, matches tables", "SQLite/Postgres", "High", 3, 1, "Planned", "T04"],
        ["T06", "Phase 1", "Career page URL finder", "Resolve career/jobs URL per company", "Node.js", "High", 6, 1, "Planned", "T04"],
        ["T07", "Phase 1", "Career page job scraper", "Extract title, location, description, apply URL", "Node.js/Cheerio", "High", 8, 1, "Planned", "T06"],
        ["T08", "Phase 1", "Job deduplication logic", "Unique by URL + company + title hash", "Node.js", "Medium", 3, 1, "Planned", "T07"],
        ["T09", "Phase 2", "Resume input & parsing", "PDF/text resume to structured sections", "Python/Node", "High", 5, 2, "Planned", "T05"],
        ["T10", "Phase 2", "Chunk resume text", "Split into skills, experience, projects, education", "Custom", "High", 2, 2, "Planned", "T09"],
        ["T11", "Phase 2", "Generate embeddings", "Ollama nomic-embed-text for resume & jobs", "Ollama API", "High", 4, 2, "Planned", "T10"],
        ["T12", "Phase 2", "Qdrant collections setup", "jobs collection + resume_chunks collection", "Qdrant", "High", 4, 2, "Planned", "T11"],
        ["T13", "Phase 2", "Similarity search", "Top 20 jobs per user resume vector", "Qdrant", "High", 4, 2, "Planned", "T12"],
        ["T14", "Phase 3", "LLM rerank prompt", "Score jobs 0-100 with fit reason & gaps", "Ollama llama3.2", "High", 4, 3, "Planned", "T13"],
        ["T15", "Phase 3", "Filter low matches", "Only send jobs with score > 70", "Custom", "Medium", 2, 3, "Planned", "T14"],
        ["T16", "Phase 4", "n8n daily cron workflow", "Schedule trigger at 8 AM daily", "n8n", "High", 3, 3, "Planned", "T07,T13"],
        ["T17", "Phase 4", "n8n scrape workflow", "Loop companies, fetch jobs, save to DB", "n8n HTTP+Code", "High", 6, 3, "Planned", "T16"],
        ["T18", "Phase 4", "n8n match workflow", "Embed new jobs, vector search, LLM rank", "n8n+Ollama", "High", 6, 4, "Planned", "T17"],
        ["T19", "Phase 5", "Email digest template", "HTML/text email with top 5-10 jobs", "Gmail SMTP", "High", 3, 4, "Planned", "T18"],
        ["T20", "Phase 5", "Telegram bot alerts", "Optional free WhatsApp alternative", "Telegram Bot API", "Medium", 3, 4, "Planned", "T18"],
        ["T21", "Phase 6", "Write README", "Setup, architecture, screenshots, how to run", "Markdown", "High", 4, 4, "Planned", "T19"],
        ["T22", "Phase 6", "Record demo video", "2-3 min walkthrough for portfolio", "Screen recorder", "Medium", 2, 4, "Planned", "T21"],
        ["T23", "Phase 6", "Resume & LinkedIn bullets", "Project description for job applications", "Docs", "High", 2, 4, "Planned", "T21"],
        ["T24", "Phase 7", "IMAP email alert parser", "Ingest Naukri/LinkedIn alert emails", "n8n IMAP", "Low", 6, "Later", "Backlog", "T18"],
        ["T25", "Phase 8", "Simple dashboard UI", "View matches, mark applied, save jobs", "React/HTML", "Low", 10, "Later", "Backlog", "T18"],
    ]

    ws.append(headers)
    for row in data:
        ws.append(row)

    style_header_row(ws, 1, len(headers))
    style_data_rows(ws, 2, len(data) + 1, len(headers))
    auto_width(ws)


def sheet_tech_stack(wb):
    ws = wb.create_sheet("Tech Stack")
    headers = ["Layer", "Tool", "Purpose", "Cost", "Notes"]
    data = [
        ["Scraping", "Node.js + Cheerio", "Fetch & parse company/job HTML", "Free", "Already built for AmbitionBox"],
        ["Scraping", "Python + BeautifulSoup", "Alternative scraper scripts", "Free", "Optional"],
        ["Automation", "n8n (self-hosted)", "Daily workflows, cron, integrations", "Free", "Docker on local machine"],
        ["LLM", "Ollama (llama3.2)", "Job ranking, explanations", "Free", "Runs locally, no API cost"],
        ["Embeddings", "Ollama (nomic-embed-text)", "Vectorize resume & job text", "Free", "768-dim embeddings"],
        ["Vector DB", "Qdrant", "Similarity search for job matching", "Free", "Docker container"],
        ["Database", "SQLite / PostgreSQL", "Companies, jobs, matches, history", "Free", "Start SQLite, migrate later"],
        ["Notifications", "Gmail SMTP", "Daily job digest email", "Free", "Personal Gmail"],
        ["Notifications", "Telegram Bot", "Instant job alerts", "Free", "Easier than WhatsApp"],
        ["Container", "Docker Compose", "Run all services together", "Free", "n8n + Ollama + Qdrant"],
        ["Version Control", "GitHub", "Portfolio repo", "Free", "Public repo with README"],
    ]

    ws.append(headers)
    for row in data:
        ws.append(row)

    style_header_row(ws, 1, len(headers))
    style_data_rows(ws, 2, len(data) + 1, len(headers))
    auto_width(ws)


def sheet_milestones(wb):
    ws = wb.create_sheet("Milestones")
    headers = ["#", "Milestone", "Target Week", "Criteria (Done When)", "Status"]
    data = [
        [1, "Company list ready", "Week 1", "CSV with 500+ companies name, slug, profile URL", "Done"],
        [2, "First jobs scraped", "Week 1", "At least 50 jobs from 10 company career pages in DB", "Planned"],
        [3, "Resume indexed", "Week 2", "Resume chunks stored with embeddings in Qdrant", "Planned"],
        [4, "First match run", "Week 2", "Top 20 jobs returned by vector similarity", "Planned"],
        [5, "LLM ranking works", "Week 3", "Ollama returns ranked jobs with fit explanation", "Planned"],
        [6, "n8n pipeline live", "Week 3", "Daily automated scrape + match without manual run", "Planned"],
        [7, "First daily alert", "Week 4", "Receive email/Telegram with top 5 matched jobs", "Planned"],
        [8, "Portfolio ready", "Week 4", "GitHub README + demo + resume bullets complete", "Planned"],
        [9, "First real application", "Week 4+", "Apply to a job found by the system", "Planned"],
        [10, "Interview mention", "Ongoing", "Explain project clearly in at least 1 interview", "Planned"],
    ]

    ws.append(headers)
    for row in data:
        ws.append(row)

    style_header_row(ws, 1, len(headers))
    style_data_rows(ws, 2, len(data) + 1, len(headers))
    auto_width(ws)


def sheet_weekly_schedule(wb):
    ws = wb.create_sheet("Weekly Schedule")
    headers = ["Week", "Mon-Tue", "Wed-Thu", "Fri-Sat", "Sun", "Goal"]
    data = [
        ["Week 1", "DB schema + career URL finder", "Career page scraper", "Test on 20 companies", "Fix bugs, CSV export", "Jobs in database"],
        ["Week 2", "Resume parser + chunking", "Ollama embeddings setup", "Qdrant indexing", "Test similarity search", "Matching works"],
        ["Week 3", "LLM rerank prompt tuning", "n8n scrape workflow", "n8n match workflow", "End-to-end test", "Automation works"],
        ["Week 4", "Email/Telegram alerts", "README + architecture doc", "Demo video", "Apply to jobs using tool", "MVP complete"],
    ]

    ws.append(headers)
    for row in data:
        ws.append(row)

    style_header_row(ws, 1, len(headers))
    style_data_rows(ws, 2, len(data) + 1, len(headers))
    auto_width(ws)


def main():
    wb = Workbook()
    sheet_overview(wb)
    sheet_phases(wb)
    sheet_tasks(wb)
    sheet_tech_stack(wb)
    sheet_milestones(wb)
    sheet_weekly_schedule(wb)
    wb.save(OUTPUT)
    print(f"Created {OUTPUT}")


if __name__ == "__main__":
    main()
